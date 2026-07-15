#!/usr/bin/env python3
"""
Generează un raport Excel pentru Irina Vilciu pe baza profilului public
Google Scholar și a criteriilor OMEC nr. 3018/2025, Comisia 9 –
Inginerie electrică.

Date extrase automat:
- publicațiile profilului Google Scholar;
- numărul de citări pentru fiecare publicație;
- lucrările citante, în limita configurată.

Date care necesită verificare manuală:
- indexarea WoS;
- indexarea IEEE Xplore;
- indexarea în celelalte baze BDI acceptate;
- calitatea de prim-autor/autor corespondent;
- dovezile verificabile.

Utilizare:
    python generate_report.py
    python generate_report.py --demo
    python generate_report.py --backend serpapi
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlparse

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


DEFAULT_PROFILE_URL = (
    "https://scholar.google.com/citations?user=sSYoPq0AAAAJ"
)
BDI_ACCEPTATE = (
    "Scopus; IEEE Xplore; ScienceDirect; Elsevier; Wiley; ACM; DBLP; "
    "SpringerLink; Engineering Village; CABI; Emerald; CSA; Compendex; "
    "IET Inspec; EBSCO; ProQuest; Index Copernicus; Ulrichsweb; DOAJ"
)


@dataclass
class Publication:
    title: str = ""
    authors: str = ""
    venue: str = ""
    year: str = ""
    link: str = ""
    citations_count: int = 0
    citations_link: str = ""
    cites_id: str = ""
    citations: list["Citation"] = field(default_factory=list)


@dataclass
class Citation:
    parent_title: str = ""
    title: str = ""
    authors: str = ""
    venue: str = ""
    year: str = ""
    link: str = ""


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "; ".join(str(item).strip() for item in value if str(item).strip())
    return str(value).strip()


def first_nonempty(*values: Any) -> str:
    for value in values:
        result = text(value)
        if result:
            return result
    return ""


def integer(value: Any) -> int:
    try:
        return int(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0


def parse_author_id(profile_url: str) -> str:
    parsed = urlparse(profile_url.strip())
    author_id = (parse_qs(parsed.query).get("user") or [""])[0].strip()
    if not author_id:
        raise ValueError(
            "URL-ul Google Scholar trebuie să conțină parametrul user."
        )
    return author_id


def build_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=4,
        read=4,
        connect=4,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        respect_retry_after_header=True,
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.headers.update(
        {
            "User-Agent": "Raportare-OMEC-3018-Comisia-9/1.0",
            "Accept": "application/json",
        }
    )
    return session


def serpapi_get(
    session: requests.Session,
    api_key: str,
    **params: Any,
) -> dict[str, Any]:
    params["api_key"] = api_key
    response = session.get(
        "https://serpapi.com/search.json",
        params=params,
        timeout=90,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("error"):
        raise RuntimeError(f"SerpApi: {payload['error']}")
    return payload


def parse_serpapi_publication(raw: dict[str, Any]) -> Publication:
    cited_by = raw.get("cited_by") or {}
    publication = first_nonempty(
        raw.get("publication"),
        (raw.get("publication_info") or {}).get("summary"),
    )
    authors = first_nonempty(
        raw.get("authors"),
        (raw.get("publication_info") or {}).get("authors"),
    )
    return Publication(
        title=first_nonempty(raw.get("title")),
        authors=authors,
        venue=publication,
        year=first_nonempty(raw.get("year")),
        link=first_nonempty(raw.get("link")),
        citations_count=integer(
            cited_by.get("value")
            or cited_by.get("total")
            or raw.get("cited_by_count")
        ),
        citations_link=first_nonempty(
            cited_by.get("link"),
            cited_by.get("serpapi_link"),
        ),
        cites_id=first_nonempty(
            cited_by.get("cites_id"),
            raw.get("cites_id"),
        ),
    )


def parse_serpapi_citation(
    parent_title: str,
    raw: dict[str, Any],
) -> Citation:
    publication_info = raw.get("publication_info") or {}
    summary = first_nonempty(publication_info.get("summary"))
    year_match = re.search(r"\b(?:19|20)\d{2}\b", summary)
    return Citation(
        parent_title=parent_title,
        title=first_nonempty(raw.get("title")),
        authors=first_nonempty(publication_info.get("authors")),
        venue=summary,
        year=year_match.group(0) if year_match else "",
        link=first_nonempty(raw.get("link")),
    )


def collect_citations_serpapi(
    session: requests.Session,
    api_key: str,
    publication: Publication,
    max_citations: int,
) -> list[Citation]:
    if not publication.cites_id or publication.citations_count <= 0:
        return []

    results: list[Citation] = []
    start = 0
    page_size = 20

    while True:
        payload = serpapi_get(
            session,
            api_key,
            engine="google_scholar",
            cites=publication.cites_id,
            start=start,
            num=page_size,
            hl="en",
        )
        organic = payload.get("organic_results") or []
        if not organic:
            break

        for raw in organic:
            results.append(parse_serpapi_citation(publication.title, raw))
            if max_citations and len(results) >= max_citations:
                return results[:max_citations]

        if len(organic) < page_size:
            break
        start += page_size
        time.sleep(0.5)

    return results


def collect_serpapi(
    author_id: str,
    api_key: str,
    max_publications: int,
    max_citations: int,
) -> tuple[dict[str, Any], list[Publication]]:
    session = build_session()
    publications: list[Publication] = []
    author_info: dict[str, Any] = {}
    start = 0
    page_size = 100

    while True:
        payload = serpapi_get(
            session,
            api_key,
            engine="google_scholar_author",
            author_id=author_id,
            start=start,
            num=page_size,
            hl="en",
        )
        if not author_info:
            author_info = payload.get("author") or {}
            author_info["cited_by"] = payload.get("cited_by") or {}

        articles = payload.get("articles") or []
        if not articles:
            break

        for raw in articles:
            publications.append(parse_serpapi_publication(raw))
            if max_publications and len(publications) >= max_publications:
                publications = publications[:max_publications]
                break

        if max_publications and len(publications) >= max_publications:
            break
        if len(articles) < page_size:
            break

        start += page_size
        time.sleep(0.5)

    for index, publication in enumerate(publications, start=1):
        print(
            f"[{index}/{len(publications)}] Citări: "
            f"{publication.title[:80]}"
        )
        publication.citations = collect_citations_serpapi(
            session,
            api_key,
            publication,
            max_citations,
        )

    return author_info, publications


def scholarly_bib_to_publication(raw: dict[str, Any]) -> Publication:
    bib = raw.get("bib") or {}
    return Publication(
        title=first_nonempty(bib.get("title")),
        authors=first_nonempty(bib.get("author")),
        venue=first_nonempty(
            bib.get("citation"),
            bib.get("journal"),
            bib.get("venue"),
            bib.get("conference"),
        ),
        year=first_nonempty(bib.get("pub_year"), bib.get("year")),
        link=first_nonempty(raw.get("pub_url"), raw.get("eprint_url")),
        citations_count=integer(raw.get("num_citations")),
        citations_link=first_nonempty(raw.get("citedby_url")),
        cites_id="",
    )


def collect_scholarly(
    author_id: str,
    max_publications: int,
    max_citations: int,
) -> tuple[dict[str, Any], list[Publication]]:
    try:
        from scholarly import scholarly
    except ImportError as exc:
        raise RuntimeError(
            "Pachetul scholarly lipsește. Rulați: "
            "python -m pip install -r requirements.txt"
        ) from exc

    author = scholarly.search_author_id(author_id)
    try:
        author = scholarly.fill(author, sections=["basics", "indices", "publications"])
    except TypeError:
        author = scholarly.fill(author)

    raw_publications = list(author.get("publications") or [])
    if max_publications:
        raw_publications = raw_publications[:max_publications]

    publications: list[Publication] = []

    for index, raw in enumerate(raw_publications, start=1):
        print(f"[{index}/{len(raw_publications)}] Publicație Google Scholar")
        try:
            filled = scholarly.fill(raw)
        except Exception:
            filled = raw

        publication = scholarly_bib_to_publication(filled)

        if publication.citations_count > 0:
            try:
                for citing_raw in scholarly.citedby(filled):
                    citing = scholarly_bib_to_publication(citing_raw)
                    publication.citations.append(
                        Citation(
                            parent_title=publication.title,
                            title=citing.title,
                            authors=citing.authors,
                            venue=citing.venue,
                            year=citing.year,
                            link=citing.link,
                        )
                    )
                    if (
                        max_citations
                        and len(publication.citations) >= max_citations
                    ):
                        break
            except Exception as exc:
                print(
                    "Avertisment: citările nu au putut fi extrase pentru "
                    f"„{publication.title}”: {exc}"
                )

        publications.append(publication)
        time.sleep(1.0)

    cited_by = author.get("citedby") or author.get("cited_by") or 0
    author_info = {
        "name": first_nonempty(author.get("name"), "Irina Vilciu"),
        "affiliations": first_nonempty(author.get("affiliation")),
        "interests": author.get("interests") or [],
        "cited_by": {"table": [{"citations": {"all": integer(cited_by)}}]},
    }
    return author_info, publications


def h_index(citation_counts: Iterable[int]) -> int:
    ordered = sorted((max(0, int(v)) for v in citation_counts), reverse=True)
    return max(
        (rank for rank, value in enumerate(ordered, start=1) if value >= rank),
        default=0,
    )


def demo_data() -> tuple[dict[str, Any], list[Publication]]:
    publication = Publication(
        title="Exemplu de publicație – înlocuit automat la rularea reală",
        authors="Irina Vilciu; Autor Exemplu",
        venue="Revistă exemplu",
        year="2025",
        link="https://scholar.google.com/",
        citations_count=1,
        citations_link="https://scholar.google.com/",
        citations=[
            Citation(
                parent_title="Exemplu de publicație – înlocuit automat la rularea reală",
                title="Exemplu de lucrare citantă",
                authors="Autor Citant",
                venue="Publicație exemplu, 2026",
                year="2026",
                link="https://scholar.google.com/",
            )
        ],
    )
    author = {
        "name": "Irina Vilciu",
        "affiliations": "Date demonstrative",
        "cited_by": {"table": [{"citations": {"all": 1}}]},
    }
    return author, [publication]


THIN_GRAY = Side(style="thin", color="B7B7B7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")


def style_header(ws, row: int, columns: int) -> None:
    for cell in ws[row][:columns]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(
            left=THIN_GRAY,
            right=THIN_GRAY,
            top=THIN_GRAY,
            bottom=THIN_GRAY,
        )


def style_data_region(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=THIN_GRAY,
                right=THIN_GRAY,
                top=THIN_GRAY,
                bottom=THIN_GRAY,
            )


def add_table(ws, reference: str, name: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def extract_author_citations(author_info: dict[str, Any]) -> int:
    cited = author_info.get("cited_by") or {}
    table = cited.get("table") or []
    if table:
        citations = (table[0].get("citations") or {}).get("all")
        return integer(citations)
    return integer(author_info.get("citedby") or author_info.get("cited_by"))


def build_workbook(
    author_info: dict[str, Any],
    publications: list[Publication],
    profile_url: str,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(
        calcMode="auto",
        fullCalcOnLoad=True,
        forceFullCalc=True,
    )

    # Liste pentru validare
    ws_lists = wb.create_sheet("Liste")
    ws_lists.append(["Stare"])
    ws_lists.append(["Da"])
    ws_lists.append(["Nu"])
    ws_lists.append(["De verificat"])
    ws_lists.sheet_state = "hidden"

    # Publicații
    ws_pub = wb.create_sheet("Publicatii")
    pub_headers = [
        "Nr.",
        "Titlu",
        "Autori",
        "Revistă / Conferință",
        "An",
        "Tip",
        "Link publicație",
        "Citări Google Scholar",
        "Link citări",
        "WoS",
        "IEEE Xplore",
        "BDI",
        "Prim-autor",
        "Autor corespondent",
        "Autor principal",
        "Criteriu WoS/IEEE",
        "Dovezi verificabile",
        "Observații",
    ]
    ws_pub.append(pub_headers)

    for index, publication in enumerate(publications, start=1):
        row = ws_pub.max_row + 1
        ws_pub.append(
            [
                index,
                publication.title,
                publication.authors,
                publication.venue,
                publication.year,
                "De verificat",
                publication.link,
                publication.citations_count,
                publication.citations_link,
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                None,
                None,
                "De verificat",
                "",
            ]
        )
        ws_pub.cell(row, 15).value = (
            f'=IF(OR(M{row}="Da",N{row}="Da"),"Da",'
            f'IF(AND(M{row}="Nu",N{row}="Nu"),"Nu","De verificat"))'
        )
        ws_pub.cell(row, 16).value = (
            f'=IF(OR(J{row}="Da",K{row}="Da"),"Da",'
            f'IF(AND(J{row}="Nu",K{row}="Nu"),"Nu","De verificat"))'
        )
        if publication.link:
            ws_pub.cell(row, 7).hyperlink = publication.link
            ws_pub.cell(row, 7).style = "Hyperlink"
        if publication.citations_link:
            ws_pub.cell(row, 9).hyperlink = publication.citations_link
            ws_pub.cell(row, 9).style = "Hyperlink"

    if not publications:
        ws_pub.append(
            [
                "",
                "Nu au fost extrase publicații.",
                "",
                "",
                "",
                "",
                "",
                0,
                "",
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                "De verificat",
                "",
            ]
        )

    style_header(ws_pub, 1, len(pub_headers))
    style_data_region(ws_pub, 2, ws_pub.max_row, len(pub_headers))
    ws_pub.freeze_panes = "A2"
    ws_pub.auto_filter.ref = f"A1:R{ws_pub.max_row}"
    if ws_pub.max_row >= 2:
        add_table(ws_pub, f"A1:R{ws_pub.max_row}", "TabelPublicatii")

    validation = DataValidation(
        type="list",
        formula1="'Liste'!$A$2:$A$4",
        allow_blank=True,
    )
    validation.error = "Selectați Da, Nu sau De verificat."
    validation.errorTitle = "Valoare neacceptată"
    ws_pub.add_data_validation(validation)
    validation.add(f"F2:F{max(ws_pub.max_row, 2)}")
    validation.add(f"J2:N{max(ws_pub.max_row, 2)}")
    validation.add(f"Q2:Q{max(ws_pub.max_row, 2)}")

    ws_pub.conditional_formatting.add(
        f"J2:Q{max(ws_pub.max_row, 2)}",
        FormulaRule(formula=['J2="Da"'], fill=GREEN_FILL),
    )
    ws_pub.conditional_formatting.add(
        f"J2:Q{max(ws_pub.max_row, 2)}",
        FormulaRule(formula=['J2="Nu"'], fill=RED_FILL),
    )
    ws_pub.conditional_formatting.add(
        f"J2:Q{max(ws_pub.max_row, 2)}",
        FormulaRule(formula=['J2="De verificat"'], fill=YELLOW_FILL),
    )

    widths = {
        "A": 7,
        "B": 48,
        "C": 34,
        "D": 36,
        "E": 10,
        "F": 16,
        "G": 34,
        "H": 18,
        "I": 34,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 16,
        "N": 20,
        "O": 18,
        "P": 19,
        "Q": 20,
        "R": 36,
    }
    for column, width in widths.items():
        ws_pub.column_dimensions[column].width = width

    # Citări
    ws_cit = wb.create_sheet("Citari")
    cit_headers = [
        "Nr.",
        "Publicația citată",
        "Titlul lucrării citante",
        "Autori",
        "Sursa",
        "An",
        "Link",
    ]
    ws_cit.append(cit_headers)
    citation_index = 0
    for publication in publications:
        for citation in publication.citations:
            citation_index += 1
            ws_cit.append(
                [
                    citation_index,
                    publication.title,
                    citation.title,
                    citation.authors,
                    citation.venue,
                    citation.year,
                    citation.link,
                ]
            )
            if citation.link:
                ws_cit.cell(ws_cit.max_row, 7).hyperlink = citation.link
                ws_cit.cell(ws_cit.max_row, 7).style = "Hyperlink"

    if citation_index == 0:
        ws_cit.append(
            [
                "",
                "",
                "Nu au fost extrase lucrări citante.",
                "",
                "",
                "",
                "",
            ]
        )

    style_header(ws_cit, 1, len(cit_headers))
    style_data_region(ws_cit, 2, ws_cit.max_row, len(cit_headers))
    ws_cit.freeze_panes = "A2"
    ws_cit.auto_filter.ref = f"A1:G{ws_cit.max_row}"
    add_table(ws_cit, f"A1:G{ws_cit.max_row}", "TabelCitari")
    for column, width in {
        "A": 7,
        "B": 45,
        "C": 48,
        "D": 34,
        "E": 38,
        "F": 10,
        "G": 34,
    }.items():
        ws_cit.column_dimensions[column].width = width

    # Criterii oficiale
    ws_crit = wb.create_sheet("Criterii_OMEC_3018")
    ws_crit["A1"] = "OMEC nr. 3018/2025 – Comisia 9: Inginerie electrică"
    ws_crit["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    ws_crit["A1"].fill = HEADER_FILL
    ws_crit.merge_cells("A1:D1")

    criteria_rows = [
        [
            "C1",
            "Publicarea a minimum două lucrări.",
            "≥ 2",
            "Evaluare automată după numărul publicațiilor extrase.",
        ],
        [
            "C2",
            "Minimum o lucrare în reviste sau publicații ale conferințelor "
            "indexate Web of Science (WoS)/IEEE Xplore.",
            "≥ 1",
            "Indexarea necesită confirmare manuală și dovadă.",
        ],
        [
            "C3",
            "Minimum o lucrare în reviste sau publicații ale conferințelor "
            "indexate într-o bază de date internațională recunoscută (BDI).",
            "≥ 1",
            "Se completează manual în foaia Publicatii.",
        ],
        [
            "C4",
            "La minimum una dintre lucrări candidatul trebuie să fie autor "
            "principal.",
            "≥ 1",
            "Autor principal = prim-autor sau autor corespondent în condițiile ordinului.",
        ],
        [
            "C5",
            "Pentru fiecare activitate declarată se prezintă dovezi verificabile.",
            "Obligatoriu",
            "Se marchează manual în foaia Publicatii.",
        ],
    ]

    ws_crit.append(["Cod", "Criteriu", "Prag", "Observații"])
    for row in criteria_rows:
        ws_crit.append(row)

    bdi_row = ws_crit.max_row + 2
    ws_crit.cell(bdi_row, 1).value = "Bazele BDI acceptate"
    ws_crit.cell(bdi_row, 1).font = Font(bold=True)
    ws_crit.cell(bdi_row, 2).value = BDI_ACCEPTATE
    ws_crit.merge_cells(
        start_row=bdi_row,
        start_column=2,
        end_row=bdi_row,
        end_column=4,
    )

    style_header(ws_crit, 2, 4)
    style_data_region(ws_crit, 3, 2 + len(criteria_rows), 4)
    ws_crit.column_dimensions["A"].width = 22
    ws_crit.column_dimensions["B"].width = 74
    ws_crit.column_dimensions["C"].width = 16
    ws_crit.column_dimensions["D"].width = 55
    ws_crit.freeze_panes = "A3"
    for row in ws_crit.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Sinteză
    ws_sum = wb.create_sheet("Sinteza", 0)
    ws_sum["A1"] = "Raportarea activității științifice – Irina Vilciu"
    ws_sum["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws_sum["A1"].fill = HEADER_FILL
    ws_sum.merge_cells("A1:D1")

    ws_sum["A3"] = "Profil Google Scholar"
    ws_sum["B3"] = profile_url
    ws_sum["B3"].hyperlink = profile_url
    ws_sum["B3"].style = "Hyperlink"
    ws_sum["A4"] = "Nume"
    ws_sum["B4"] = first_nonempty(author_info.get("name"), "Irina Vilciu")
    ws_sum["A5"] = "Afiliere"
    ws_sum["B5"] = first_nonempty(
        author_info.get("affiliations"),
        author_info.get("affiliation"),
    )
    ws_sum["A6"] = "Data generării"
    ws_sum["B6"] = time.strftime("%Y-%m-%d %H:%M:%S")

    ws_sum["A8"] = "Indicator"
    ws_sum["B8"] = "Valoare"
    ws_sum["C8"] = "Prag"
    ws_sum["D8"] = "Evaluare"

    last_pub_row = ws_pub.max_row
    ws_sum["A9"] = "Număr publicații"
    ws_sum["B9"] = max(len(publications), 0)
    ws_sum["C9"] = 2
    ws_sum["D9"] = '=IF(B9>=C9,"Îndeplinit","Neîndeplinit")'

    ws_sum["A10"] = "Citări totale în publicațiile extrase"
    ws_sum["B10"] = sum(p.citations_count for p in publications)
    ws_sum["C10"] = "Informativ"
    ws_sum["D10"] = "Indicator bibliometric"

    ws_sum["A11"] = "Indice Hirsch calculat din lista extrasă"
    ws_sum["B11"] = h_index(p.citations_count for p in publications)
    ws_sum["C11"] = "Informativ"
    ws_sum["D11"] = "Poate diferi dacă extragerea este limitată"

    ws_sum["A12"] = "Lucrări WoS/IEEE Xplore confirmate"
    ws_sum["B12"] = f'=COUNTIF(Publicatii!P2:P{last_pub_row},"Da")'
    ws_sum["C12"] = 1
    ws_sum["D12"] = '=IF(B12>=C12,"Îndeplinit",IF(COUNTIF(Publicatii!P2:P1048576,"De verificat")>0,"Necesită verificare","Neîndeplinit"))'

    ws_sum["A13"] = "Lucrări BDI confirmate"
    ws_sum["B13"] = f'=COUNTIF(Publicatii!L2:L{last_pub_row},"Da")'
    ws_sum["C13"] = 1
    ws_sum["D13"] = '=IF(B13>=C13,"Îndeplinit",IF(COUNTIF(Publicatii!L2:L1048576,"De verificat")>0,"Necesită verificare","Neîndeplinit"))'

    ws_sum["A14"] = "Lucrări ca autor principal"
    ws_sum["B14"] = f'=COUNTIF(Publicatii!O2:O{last_pub_row},"Da")'
    ws_sum["C14"] = 1
    ws_sum["D14"] = '=IF(B14>=C14,"Îndeplinit",IF(COUNTIF(Publicatii!O2:O1048576,"De verificat")>0,"Necesită verificare","Neîndeplinit"))'

    ws_sum["A15"] = "Lucrări cu dovezi verificabile"
    ws_sum["B15"] = f'=COUNTIF(Publicatii!Q2:Q{last_pub_row},"Da")'
    ws_sum["C15"] = "Pentru fiecare activitate declarată"
    ws_sum["D15"] = '=IF(B15=B9,"Îndeplinit","Necesită verificare")'

    ws_sum["A17"] = "Verdict preliminar"
    ws_sum["B17"] = (
        '=IF(AND(B9>=2,B12>=1,B13>=1,B14>=1,B15=B9),'
        '"Criterii minimale îndeplinite preliminar",'
        '"Necesită completare și verificare manuală")'
    )
    ws_sum.merge_cells("B17:D17")

    style_header(ws_sum, 8, 4)
    style_data_region(ws_sum, 9, 15, 4)
    for cell in ("A3", "A4", "A5", "A6", "A17"):
        ws_sum[cell].font = Font(bold=True)
        ws_sum[cell].fill = SECTION_FILL
    ws_sum["B17"].font = Font(bold=True)
    ws_sum["B17"].alignment = Alignment(wrap_text=True)

    ws_sum.column_dimensions["A"].width = 42
    ws_sum.column_dimensions["B"].width = 34
    ws_sum.column_dimensions["C"].width = 30
    ws_sum.column_dimensions["D"].width = 34
    ws_sum.freeze_panes = "A8"

    for row in ws_sum.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.properties.title = "Raport OMEC 3018/2025 – Irina Vilciu"
    wb.properties.subject = "Comisia 9 – Inginerie electrică"
    wb.properties.creator = "Raportare GitHub"
    wb.save(output_path)


def positive_limit(value: str | int | None) -> int:
    result = integer(value)
    return max(result, 0)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extrage publicațiile și citările Google Scholar și generează "
            "raportul Excel OMEC 3018/2025 – Comisia 9."
        )
    )
    parser.add_argument(
        "--backend",
        choices=("auto", "serpapi", "scholarly"),
        default=None,
        help="Metoda de acces Google Scholar.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Calea fișierului Excel rezultat.",
    )
    parser.add_argument(
        "--max-publications",
        type=int,
        default=None,
        help="0 = fără limită.",
    )
    parser.add_argument(
        "--max-citations",
        type=int,
        default=None,
        help="Numărul maxim de citări extrase pentru fiecare publicație; 0 = fără limită.",
    )
    parser.add_argument(
        "--demo",
        action="store_true",
        help="Generează un raport demonstrativ fără acces la internet.",
    )
    return parser.parse_args()


def main() -> int:
    load_dotenv()
    args = parse_arguments()

    profile_url = os.getenv(
        "GOOGLE_SCHOLAR_PROFILE_URL",
        DEFAULT_PROFILE_URL,
    ).strip()
    author_id = parse_author_id(profile_url)

    backend = (
        args.backend
        or os.getenv("SCHOLAR_BACKEND", "auto").strip().lower()
        or "auto"
    )
    api_key = os.getenv("SERPAPI_API_KEY", "").strip()

    max_publications = (
        positive_limit(args.max_publications)
        if args.max_publications is not None
        else positive_limit(os.getenv("MAX_PUBLICATIONS", "0"))
    )
    max_citations = (
        positive_limit(args.max_citations)
        if args.max_citations is not None
        else positive_limit(
            os.getenv("MAX_CITATIONS_PER_PUBLICATION", "0")
        )
    )

    output_value = (
        args.output
        or os.getenv(
            "OUTPUT_FILE",
            "data/output/Raport_OMEC_3018_Comisia_9_Irina_Vilciu.xlsx",
        )
    )
    output_path = Path(output_value).expanduser()

    if args.demo:
        author_info, publications = demo_data()
        active_backend = "demo"
    else:
        if backend == "auto":
            active_backend = "serpapi" if api_key else "scholarly"
        else:
            active_backend = backend

        print(f"Profil: {profile_url}")
        print(f"Author ID: {author_id}")
        print(f"Backend: {active_backend}")

        if active_backend == "serpapi":
            if not api_key:
                raise RuntimeError(
                    "Lipsește SERPAPI_API_KEY în fișierul local .env."
                )
            author_info, publications = collect_serpapi(
                author_id,
                api_key,
                max_publications,
                max_citations,
            )
        else:
            author_info, publications = collect_scholarly(
                author_id,
                max_publications,
                max_citations,
            )

    build_workbook(author_info, publications, profile_url, output_path)

    print()
    print("Raport generat cu succes:")
    print(output_path.resolve())
    print(f"Publicații: {len(publications)}")
    print(
        "Lucrări citante extrase: "
        f"{sum(len(p.citations) for p in publications)}"
    )
    print(
        "Completați manual coloanele WoS, IEEE Xplore, BDI, "
        "autor principal și dovezi verificabile."
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nOperație întreruptă de utilizator.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as exc:
        print(f"\nEROARE: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise SystemExit(1)
